# hint: you can get a lot of macro informations in your docs using:
#   {{ macros_info() }}


def define_env(env):

    # ==================================================================
    @env.macro
    def lorem():
        """Just print lorem ipsum."""
        return """Lorem ipsum dolor sit amet, consectetur adipisicing elit, sed do eiusmod
tempor incididunt ut labore et dolore magna aliqua. Ut enim ad minim veniam,
quis nostrud exercitation ullamco laboris nisi ut aliquip ex ea commodo
consequat. Duis aute irure dolor in reprehenderit in voluptate velit esse
cillum dolore eu fugiat nulla pariatur. Excepteur sint occaecat cupidatat non
proident, sunt in culpa qui officia deserunt mollit anim id est laborum.
"""

    # ==================================================================
    @env.macro
    def csv2table(filename, dialect="excel"):
        """
        Converts a csv file living relative to the current document
        to a html table.
        """

        import csv

        try:

            abs_filename = "/".join(env.page.file.abs_src_path.split("/")[0:-1] + [filename])
            
            header = None
            lines = []

            with open(abs_filename, newline="") as csv_file:
                my_reader = csv.reader(csv_file, dialect=dialect)
                for row in my_reader:
                    if not header:
                        header = row
                    else:
                        lines.append(row)

            header_html = "".join(f"<th>{s}</th>" for s in header)
            lines_html = ""
            for line in lines:
                lines_html += "<tr>" + "".join(f"<td>{s.strip()}</td>" for s in line) + "</tr>"

            return f"<table><thead><tr>{header_html}</tr></thead><tbody>{lines_html}</tbody></table>"

        except:
            return f"""<p style="color: black; background: red; padding: 1rem;"><b>Error reading CSV file {filename}.</b></p>"""

    # ==================================================================

    @env.macro
    def xlsx2table(filename):
        """
        Convert an excel file living relative to the current document
        to a html table.
        """

        try:
            import openpyxl
            from pathlib import Path

            abs_filename = "/".join(env.page.file.abs_src_path.split("/")[0:-1] + [filename])

            xlsx_file = Path(abs_filename)
            workbook = openpyxl.load_workbook(xlsx_file) 

            sheet = workbook.active

            # max_row = sheet.max_row
            max_column = sheet.max_column

            header = None
            lines = []

            for row in sheet.iter_rows():
                if not header:
                    header = row
                else:
                    lines.append(row)

            header_html = "".join(f"<th>{s.value}</th>" for s in header[0:max_column])
            lines_html = ""
            for line in lines:
                lines_html += "<tr>" + "".join(f"<td>{s.value}</td>" for s in line[0:max_column]) + "</tr>"

            return f"<table><thead><tr>{header_html}</tr></thead><tbody>{lines_html}</tbody></table>"

        except:
            return f"""<p style="color: black; background: red; padding: 1rem;"><b>Error reading XLSX file {filename}.</b></p>"""

    # ==================================================================

    @env.macro
    def begin_page():
        """
        Opens a div.page the will be closed with a end().
        There are some css rules to let the browser try to avoid page
        breaks inside this div.page and put one behind, if need be.
        """
        return '<div class="page" markdown="1">'


    @env.macro
    def end_page():
        """See begin()."""
        return '</div><!-- .page -->'

